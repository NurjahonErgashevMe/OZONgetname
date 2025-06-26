import logging
import threading
import queue
import time
import re
import os
import openpyxl
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import undetected_chromedriver as uc
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from config import *

class OzonProductParser:
    def __init__(self, category_name):
        self.results = []
        self.processed_count = 0
        self.stop_event = threading.Event()
        self.logger = logging.getLogger('product_parser')
        self.workbook = None
        self.worksheet = None
        self.drivers = []
        self.category_name = category_name
        self.timestamp = get_timestamp()
        
        # Формируем имя файла
        self.excel_filename = os.path.join(
            RESULTS_DIR, 
            f"{self.category_name}_{self.timestamp}.xlsx"
        )
        self.logger.info(f"Файл результатов: {self.excel_filename}")

    def init_workbook(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Результаты парсинга Ozon"

    def create_driver(self):
        options = uc.ChromeOptions()
        options.add_argument('--disable-gpu')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--blink-settings=imagesEnabled=false')
        
        prefs = {'profile.default_content_setting_values': {'images': 2, 'javascript': 1}}
        options.add_experimental_option('prefs', prefs)
        
        driver = uc.Chrome(
            options=options,
            driver_executable_path=CHROMEDRIVER_PATH,
            headless=False,
            use_subprocess=True
        )
        self.drivers.append(driver)
        return driver

    def parse_page(self, driver, url):
        result = {'product': 'Not Found', 'seller': 'Not Found', 'status': 'success'}
        
        try:
            driver.get(url)
            
            # Проверка наличия товара
            try:
                WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((By.XPATH, '//div[@data-widget="webOutOfStock"]'))
                )
                result['status'] = "out_of_stock"
                
                # Парсинг для отсутствующего товара
                try:
                    product_element = driver.find_element(
                        By.XPATH, 
                        '//div[@data-widget="webOutOfStock"]//p[contains(@class, "yl6_27")]'
                    )
                    result['product'] = product_element.text.strip()
                except:
                    try:
                        product_element = driver.find_element(
                            By.XPATH, 
                            '//div[@data-widget="webOutOfStock"]//p[contains(text(), "Intel") or contains(text(), "Системный")]'
                        )
                        result['product'] = product_element.text.strip()
                    except:
                        result['product'] = "Название не найдено"
                
                try:
                    seller_element = driver.find_element(
                        By.XPATH, 
                        '//div[@data-widget="webOutOfStock"]//a[contains(@href, "/seller/")]'
                    )
                    result['seller'] = seller_element.text.strip()
                except:
                    result['seller'] = "Продавец не найден"
                    
                return result
            except TimeoutException:
                pass
            
            # Поиск названия товара
            try:
                product_element = WebDriverWait(driver, 10).until(
                    EC.visibility_of_element_located((
                        By.XPATH, 
                        '//div[@data-widget="webProductHeading"]//h1 | '
                        '//div[@data-widget="webProductHeading"]//*[contains(@class, "m9p_27")] | '
                        '//h1[@data-widget="webProductHeading"]'
                    ))
                )
                result['product'] = product_element.text.strip()
                
            except TimeoutException:
                try:
                    product_container = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((
                            By.XPATH, 
                            '//div[contains(@class, "p9m_27")]'
                        ))
                    )
                    product_element = product_container.find_element(By.TAG_NAME, 'h1')
                    result['product'] = product_element.text.strip()
                    
                except:
                    try:
                        product_js = driver.execute_script(
                            """
                            const widget = document.querySelector('div[data-widget="webProductHeading"]');
                            if (widget) {
                                const h1 = widget.querySelector('h1');
                                if (h1) return h1.innerText.trim();
                                return widget.innerText.trim();
                            }
                            
                            const classElement = document.querySelector('.m9p_27, .tsHeadline');
                            if (classElement) return classElement.innerText.trim();
                            
                            return document.title.split('|')[0].trim();
                            """
                        )
                        if product_js:
                            result['product'] = product_js
                        else:
                            result['product'] = "Название не найдено"
                    except:
                        result['product'] = "Название не найдено"
            
            # Скролл для активации контента
            driver.execute_script("window.scrollTo(0, 600);")
            time.sleep(0.5)
            
            # Поиск продавца
            try:
                seller_element = WebDriverWait(driver, 5).until(
                    EC.visibility_of_element_located((
                        By.XPATH, 
                        '//div[@data-widget="webCurrentSeller"]//a[contains(@class, "s1k_27") and @title] | '
                        '//a[contains(@href, "/seller/") and @title]'
                    ))
                )
                result['seller'] = seller_element.text.strip()
            except TimeoutException:
                seller_locators = [
                    (By.XPATH, '//div[@data-widget="webCurrentSeller"]//a[@title]'),
                    (By.CSS_SELECTOR, 'div[data-widget="webCurrentSeller"] a.s1k_27'),
                    (By.XPATH, '//a[contains(@class, "s1k_27") and @title]'),
                    (By.XPATH, '//a[contains(@href, "/seller/")]')
                ]
                
                for locator in seller_locators:
                    try:
                        element = WebDriverWait(driver, 2).until(
                            EC.visibility_of_element_located(locator)
                        )
                        result['seller'] = element.text.strip()
                        break
                    except TimeoutException:
                        continue
                else:
                    try:
                        seller_js = driver.execute_script(
                            """
                            const sellerElement = 
                                document.querySelector('div[data-widget="webCurrentSeller"] a.s1k_27[title]') ||
                                document.querySelector('div[data-widget="webCurrentSeller"] a[title]') ||
                                document.querySelector('a.s1k_27[title]') ||
                                document.querySelector('a[href*="/seller/"][title]');
                            return sellerElement ? sellerElement.textContent.trim() : '';
                            """
                        )
                        if seller_js:
                            result['seller'] = seller_js
                        else:
                            result['seller'] = "Продавец не найден"
                    except:
                        result['seller'] = "Продавец не найден"
                
        except Exception as e:
            result['status'] = "error"
            result['seller'] = f"Ошибка: {str(e)}"
            
        return result

    def worker(self, url_queue, worker_id):
        driver = self.create_driver()
        self.logger.info(f"Воркер {worker_id} запущен")
        
        while not url_queue.empty() and not self.stop_event.is_set():
            try:
                url = url_queue.get_nowait()
                
                result = self.parse_page(driver, url)
                
                with threading.Lock():
                    self.results.append((url, result['product'], result['seller'], result['status']))
                    self.processed_count += 1
                    current_count = self.processed_count
                
                status_msg = {
                    "out_of_stock": "ЗАКОНЧИЛСЯ",
                    "error": "ОШИБКА",
                    "success": "УСПЕШНО"
                }.get(result['status'], "НЕИЗВЕСТНО")
                
                self.logger.info(f"[{current_count}/{self.total_urls}] {status_msg}: {url}")
                self.logger.info(f"   Товар: {result['product']}")
                self.logger.info(f"   Продавец: {result['seller']}")
                
            except queue.Empty:
                break
            finally:
                url_queue.task_done()
                time.sleep(0.3)
                
        driver.quit()
        self.logger.info(f"Воркер {worker_id} завершил работу")

    def clean_text_value(self, value):
        if not value:
            return ""
        return str(value).replace('\n', ' ').replace('\r', ' ').strip()

    def save_results(self):
        try:
            self.init_workbook()
            ws = self.worksheet
            
            headers = ['URL товара', 'Название товара', 'Продавец']
            
            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            data_font = Font(name='Arial', size=11)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            status_colors = {
                'success': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                'out_of_stock': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                'error': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                'not_found': PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid'),
                'seller_not_found': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            }
            
            # Заголовки
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Данные
            for row_num, (url, product, seller, status) in enumerate(self.results, 2):
                clean_product = self.clean_text_value(product)
                clean_seller = self.clean_text_value(seller)
                
                row_data = [url, clean_product, clean_seller]
                
                if status == "success" and "не найден" in clean_seller.lower():
                    row_fill = status_colors['seller_not_found']
                else:
                    row_fill = status_colors.get(status, None)
                
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
                    
                    if row_fill:
                        cell.fill = row_fill
            
            # Настройки столбцов
            column_widths = [75, 45, 45]
            for col_num, width in enumerate(column_widths, 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = width
            
            # Высота строк
            for row in range(1, len(self.results) + 2):
                ws.row_dimensions[row].height = 25
            
            # Автофильтр и закрепление
            if self.results:
                ws.auto_filter.ref = f"A1:C{len(self.results) + 1}"
            ws.freeze_panes = "A2"
            
            # Сохранение
            self.workbook.save(self.excel_filename)
            
            self.logger.info(f"Результаты сохранены в {self.excel_filename}")
            return True
            
        except Exception as e:
            self.logger.error(f"Ошибка при сохранении Excel: {str(e)}")
            return False

    def close_all_drivers(self):
        for driver in self.drivers:
            try:
                driver.quit()
            except:
                pass
        self.logger.info("Все браузеры закрыты")

    def run(self, urls):
        self.total_urls = len(urls)
        if not urls:
            self.logger.error("Нет URL для обработки")
            return False
        
        start_time = time.time()
        self.logger.info(f"Начало обработки {self.total_urls} товаров")
        
        # Создание очереди задач
        url_queue = queue.Queue()
        for url in urls:
            url_queue.put(url)
        
        # Запуск воркеров
        workers = []
        for i in range(WORKER_COUNT):
            worker_thread = threading.Thread(
                target=self.worker,
                args=(url_queue, i+1),
                daemon=True
            )
            worker_thread.start()
            workers.append(worker_thread)
            time.sleep(1)
        
        # Ожидание завершения
        try:
            url_queue.join()
        except KeyboardInterrupt:
            self.stop_event.set()
            self.logger.warning("Получен сигнал прерывания!")
        
        # Финализация
        success = self.save_results()
        duration = time.time() - start_time
        
        if success:
            self.logger.info(f"Обработка завершена за {duration:.2f} секунд")
            self.logger.info(f"Средняя скорость: {self.total_urls/duration:.2f} товаров/сек")
        
        # Закрываем все браузеры
        self.close_all_drivers()
        
        return success