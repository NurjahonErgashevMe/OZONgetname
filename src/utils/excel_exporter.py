import os
import logging
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from src.config import RESULTS_DIR

class ExcelExporter:
    def __init__(self, category_name, timestamp):
        self.logger = logging.getLogger('excel_exporter')
        self.category_name = category_name
        self.timestamp = timestamp
        self.workbook = None
        self.worksheet = None

        self.excel_filename = os.path.join(
            RESULTS_DIR, 
            f"{self.category_name}_{self.timestamp}.xlsx"
        )
        self.logger.info(f"Файл результатов: {self.excel_filename}")

    def init_workbook(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Результаты парсинга Ozon"

    def save_results(self, results):
        try:
            self.init_workbook()
            ws = self.worksheet

            headers = ['Название товара', 'Название компании', 'Ссылка на товар', 'Ссылка на изображение']

            header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            data_font = Font(name='Arial', size=11)
            data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            status_colors = {
                'success': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
                'out_of_stock': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
                'error': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
                'not_found': PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid'),
                'seller_not_found': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
            }

            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for row_num, result in enumerate(results, 2):
                if isinstance(result, dict):
                    url = result.get('url', '')
                    product = result.get('product_name', '')
                    company = result.get('company_name', '')
                    image_url = result.get('image_url', '')
                    status = result.get('status', 'success')
                else:
                    if len(result) >= 3:
                        url, product, company = result[0], result[1], result[2]
                        image_url = result[3] if len(result) >= 4 else "Не найдено"
                        status = result[5] if len(result) >= 6 else "success"
                    else:
                        url = str(result[0]) if len(result) > 0 else ""
                        product = str(result[1]) if len(result) > 1 else "Не найдено"
                        company = str(result[2]) if len(result) > 2 else "Не найдено"
                        image_url = "Не найдено"
                        status = "success"

                clean_product = self._clean_text_value(product)
                clean_company = self._clean_text_value(company)

                row_data = [clean_product, clean_company, url, image_url]

                row_fill = status_colors.get(status, None)

                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border

                    if row_fill:
                        cell.fill = row_fill

            column_widths = [60, 40, 75, 75]
            for col_num, width in enumerate(column_widths, 1):
                col_letter = get_column_letter(col_num)
                ws.column_dimensions[col_letter].width = width

            for row in range(1, len(results) + 2):
                ws.row_dimensions[row].height = 25

            if results:
                ws.auto_filter.ref = f"A1:D{len(results) + 1}"
            ws.freeze_panes = "A2"

            self.workbook.save(self.excel_filename)
            self.logger.info(f"Результаты сохранены в {self.excel_filename}")
            return True

        except Exception as e:
            self.logger.error(f"Ошибка при сохранении Excel: {str(e)}")
            return False

    def _clean_text_value(self, value):
        if not value:
            return ""
        return str(value).replace('\n', ' ').replace('\r', ' ').strip()

    def get_filename(self):
        return self.excel_filename
